from flask import Blueprint, send_file, flash, redirect, url_for, request, jsonify, current_app
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import subprocess
import shutil
import zipfile
import json
import threading
import time
import schedule


from models import assets_collection
from init_db import asset_type_fields

export_bp = Blueprint('export', __name__)

# === 📥 1. EXPORT KEKA =======================================
@export_bp.route('/keka')
def export_keka():
    assets = list(assets_collection.find())

    # Define column headers
    headers = [
        "Asset ID", "Asset Name", "Asset Description", "Asset Location", "Asset Category",
        "Asset Type", "Purchased On (dd-mmm-yyyy)", "Warranty Expires On (dd-mmm-yyyy)",
        "Asset Condition", "Asset Status", "Reason, if Not Available",
        "Employee Number, if Assigned", "Date of Asset Assignment (dd-mmm-yyyy)"
    ]

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "KEKA Export"

    # Apply header styles
    bold_font = Font(bold=True, name="Calibri")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold_font
            cell.alignment = wrap_align
            ws.column_dimensions[cell.column_letter].width = 25  # Adjust width as needed

    for asset in assets:
        # Asset ID: first available from the list
        id_fields = ["asset_tag", "endpoint_name", "serial_no", "mtr_asset_tag", "monitor_asset_tag", "cpu_asset_tag"]
        asset_id = next((asset.get(f, "").strip() for f in id_fields if asset.get(f)), "")

        # Asset Name & Type: category
        asset_type = asset.get("category", "").strip()

        # Description: model + system_model + ram + storage
        desc_parts = [asset.get("model", "").strip(), asset.get("system_model", "").strip(), asset.get("ram", ""). strip(), asset.get("storage", ""). strip()]
        asset_desc = "  ".join([d for d in desc_parts if d])

        #description: model + system_model
        # Location: area + (state)
        area = asset.get("area", "").strip()
        state = asset.get("state", "").strip()
        location = f"{area} ({state})" if state else area

        # Static Asset Category
        asset_category = "IT assets"

        # Purchased On & Assignment Date: from given_date
        def format_date(d):
            try:
                return datetime.strptime(d, "%d-%m-%Y").strftime("%d-%b-%Y")
            except Exception:
                return ""

        given_date = format_date(asset.get("given_date", "").strip())

        # Warranty Expires On — blank for now
        warranty_expires = ""

        # Status fields
        status = asset.get("status", "").strip()

        # Employee Number: username (user_code)
        username = asset.get("username", "").strip()
        user_code = asset.get("user_code", "").strip()
        #employee_number = f"{username} ({user_code})" if username or user_code else ""
        if user_code:
            employee_number = f"{username} ({user_code})"
        else:
            employee_number = username

        # Append row
        ws.append([
            asset_id,
            asset_type,
            asset_desc,
            location,
            asset_category,
            asset_type,
            given_date,
            warranty_expires,
            status,
            status,
            status,
            employee_number,
            given_date
        ])

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"KEKA_Asset_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# === 📥 2. EXPORT EXCEL =======================================
@export_bp.route('/excel')
def export_excel():
    assets = list(assets_collection.find())

    # Group assets by type (category)
    assets_by_type = {}
    for asset in assets:
        asset_type = asset.get("category", "Unknown").strip()
        assets_by_type.setdefault(asset_type, []).append(asset)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for asset_type, asset_list in assets_by_type.items():
        fields = asset_type_fields.get(asset_type)
        if not fields:
            continue  # Skip unknown types not in asset_type_fields

        sheet = wb.create_sheet(title=asset_type)
        headers = [field["label"] for field in fields]
        keys = [field["name"] for field in fields]
        types = [field.get("type", "text") for field in fields]

        # Styles
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell_font = Font(name="Calibri")
        fill = PatternFill(start_color="043251", end_color="043251", fill_type="solid")
        align_wrap = Alignment(wrap_text=True, vertical="top")

        # Write header row
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = align_wrap
            sheet.column_dimensions[cell.column_letter].width = 25

        # Write data rows
        for row_idx, asset in enumerate(asset_list, start=2):
            for col_idx, (key, dtype) in enumerate(zip(keys, types), start=1):
                raw_value = asset.get(key, "")

                # Formatting
                if dtype == "date" and isinstance(raw_value, str):
                    try:
                        date_obj = datetime.strptime(raw_value.strip(), "%d-%m-%Y")
                        value = date_obj.strftime("%d-%m-%Y")
                    except Exception:
                        value = raw_value
                elif dtype == "currency":
                    try:
                        value = f"₹{float(raw_value):,.2f}"
                    except Exception:
                        value = raw_value
                else:
                    value = str(raw_value)

                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = align_wrap

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Asset_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# === 📤 3. EXPORT MONGODB DATABASE =======================================
@export_bp.route('/export_db')
def export_db():
    BACKUP_FOLDER = 'mongo_backups'
    DB_NAME = 'ams'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dump_path = os.path.join(BACKUP_FOLDER, f"{DB_NAME}_{timestamp}")

    try:
        os.makedirs(BACKUP_FOLDER, exist_ok=True)

        # Clean old backup
        if os.path.exists(dump_path):
            shutil.rmtree(dump_path)

        # Run mongodump
        MONGODUMP_PATH = r"C:\Users\Admin\Downloads\mongodb-tools\bin\mongodump.exe"
        subprocess.run([MONGODUMP_PATH, '--db', DB_NAME, '--out', BACKUP_FOLDER], check=True)


        flash('✅ MongoDB export completed successfully.', 'success')
    except Exception as e:
        flash(f'❌ Export failed: {e}', 'danger')

    return redirect(url_for('main.dashboard'))

# === 📥 4. IMPORT EXCEL =======================================
@export_bp.route('/import_excel')
def import_excel():
    flash("📥 Import Excel coming soon!", "info")
    return redirect(url_for('main.dashboard'))


# === 📥 5. IMPORT MONGODB DATABASE =======================================
@export_bp.route('/import_db')
def import_db():
    BACKUP_FOLDER = 'mongo_backups'
    DB_NAME = 'ams'
    restore_path = os.path.join(BACKUP_FOLDER, DB_NAME)

    try:
        folders = sorted(
            [f for f in os.listdir(BACKUP_FOLDER) if f.startswith(DB_NAME)],
            reverse=True
        )
        if not folders:
            flash('⚠️ No backup folders found.', 'warning')
            return redirect(url_for('dashboard'))

        latest = folders[0]
        restore_path = os.path.join(BACKUP_FOLDER, latest)

        MONGORESTORE_PATH = r"C:\Users\Admin\Downloads\mongodb-tools\bin\mongorestore.exe"
        subprocess.run([MONGORESTORE_PATH, '--drop', '--db', DB_NAME, restore_path], check=True)


        flash('✅ MongoDB import completed successfully.', 'success')
    except Exception as e:
        flash(f'❌ Import failed: {e}', 'danger')

    return redirect(url_for('main.dashboard'))

def run_weekly_backup():
    BACKUP_FOLDER = 'mongo_backups'
    DB_NAME = 'ams'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dump_path = os.path.join(BACKUP_FOLDER, f"{DB_NAME}_{timestamp}")

    try:
        os.makedirs(BACKUP_FOLDER, exist_ok=True)
        subprocess.run([
            r"C:\Users\Admin\Downloads\mongodb-tools\bin\mongodump.exe",
            '--db', DB_NAME, '--out', BACKUP_FOLDER
        ], check=True)
        print(f'✅ Weekly MongoDB backup completed at {timestamp}.')
    except Exception as e:
        print(f'❌ Weekly backup failed: {e}')


def start_backup_scheduler():
    schedule.every().week.do(run_weekly_backup)

    def run():
        while True:
            schedule.run_pending()
            time.sleep(60)

    threading.Thread(target=run, daemon=True).start()

@export_bp.route('/manual_backup')
def manual_backup():
    try:
        run_weekly_backup()
        flash("✅ Manual MongoDB backup completed successfully.", "success")
    except Exception as e:
        flash(f"❌ Manual backup failed: {e}", "danger")
    return redirect(url_for('main.dashboard'))


start_backup_scheduler()
